from flask import Flask, request, send_file, send_from_directory
from flask_cors import CORS
import os
import re
import time
from PyPDF2 import PdfReader
import openpyxl
from openpyxl.utils import get_column_letter

# Use temporary storage for Render
UPLOAD_DIR = "/tmp/uploads"
OUTPUT_DIR = "/tmp/outputs"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

app = Flask(__name__, static_folder="frontend", static_url_path="/static")
CORS(app)

REG_COL = "B"
START_ROW = 8
SUBJECT_ROW = 5


@app.route("/")
def home():
    return send_from_directory("frontend", "index.html")


def normalize(code):
    code = str(code).upper()
    return re.sub(r"[^A-Z0-9]", "", code)


def detect_subjects(ws):
    subject_map = {}

    for col in range(1, ws.max_column):
        cell = ws.cell(row=SUBJECT_ROW, column=col).value
        if not cell:
            continue

        raw = str(cell).upper()
        parts = re.split(r"[\/,]", raw)

        for p in parts:
            code = normalize(p)
            if not code:
                continue

            subject_map[code] = {
                "EXT": get_column_letter(col),
                "INT": get_column_letter(col + 1),
                "TOT": get_column_letter(col + 2),
            }

    return subject_map


def find_row(ws, reg):
    reg = str(reg).strip()

    for r in range(START_ROW, ws.max_row + 1):
        val = ws[f"{REG_COL}{r}"].value

        if val:
            try:
                val = str(int(float(val)))
            except:
                val = str(val)

            if val == reg:
                return r

    return None


def to_int(v):
    try:
        return int(str(v).strip())
    except:
        return 0


@app.route("/convert", methods=["POST"])
def convert():
    pdf = request.files.get("pdf")
    template = request.files.get("template")
    sheet = request.form.get("sheet_name")

    if not pdf or not template:
        return "Missing PDF or Excel", 400

    ts = int(time.time())

    pdf_path = os.path.join(UPLOAD_DIR, f"in_{ts}.pdf")
    tmp_path = os.path.join(UPLOAD_DIR, f"tmp_{ts}.xlsx")
    out_path = os.path.join(OUTPUT_DIR, f"out_{ts}.xlsx")

    pdf.save(pdf_path)
    template.save(tmp_path)

    wb = openpyxl.load_workbook(tmp_path)

    if sheet not in wb.sheetnames:
        return "Sheet not found", 400

    ws = wb[sheet]
    SUBJECTS = detect_subjects(ws)

    reader = PdfReader(pdf_path)

    reg_re = re.compile(r"Register Number\s*:\s*(\d+)", re.I)

    subject_re = re.compile(
        r"\b([A-Z0-9]{4,10})\b\s+"
        r"(\d+)\s+"
        r"(\d+|---|-)\s+"
        r"(\d+|\*\*\*|RA|RK|---|-)",
        re.I
    )

    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue

        text = re.sub(r"\s+", " ", text)

        reg_match = reg_re.search(text)
        if not reg_match:
            continue

        regno = reg_match.group(1)
        row = find_row(ws, regno)

        if not row:
            continue

        subjects = subject_re.findall(text)

        for code, ue, ia, tot in subjects:
            code = normalize(code)

            if code not in SUBJECTS:
                continue

            ue = to_int(ue)

            if ia in ["---", "-"]:
                ia = 0
            else:
                ia = to_int(ia)

            if tot in ["***", "RA", "RK", "---", "-"]:
                tot = ue + ia
            else:
                tot = to_int(tot)

            cols = SUBJECTS[code]

            ws[f"{cols['EXT']}{row}"] = ue
            ws[f"{cols['INT']}{row}"] = ia
            ws[f"{cols['TOT']}{row}"] = tot

    wb.save(out_path)

    return send_file(
        out_path,
        as_attachment=True,
        download_name="college_filled_marks.xlsx"
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)